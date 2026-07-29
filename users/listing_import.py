"""Bulk listing import from CSV, Excel, or Google Sheets."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import BinaryIO
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, urlopen

from django.core.exceptions import ValidationError
from django.core.validators import URLValidator

from users.industries import industry_choices_flat
from users.models import Business


IMPORT_COLUMNS = [
    "name",
    "industry_type",
    "public_phone",
    "public_email",
    "public_address",
    "hero_title",
    "hero_subtitle",
    "hero_image_url",
    "website_url",
    "map_embed_url",
    "testimonial_quote",
    "testimonial_author",
    "listing_plan",
    "timezone",
    "upi_id",
    "slug",
]

# Friendly header aliases → canonical column
HEADER_ALIASES = {
    "name": "name",
    "business name": "name",
    "business_name": "name",
    "listing name": "name",
    "title": "name",
    "industry_type": "industry_type",
    "industry": "industry_type",
    "category": "industry_type",
    "public_phone": "public_phone",
    "phone": "public_phone",
    "mobile": "public_phone",
    "contact": "public_phone",
    "public_email": "public_email",
    "email": "public_email",
    "public_address": "public_address",
    "address": "public_address",
    "location": "public_address",
    "hero_title": "hero_title",
    "headline": "hero_title",
    "hero_subtitle": "hero_subtitle",
    "subtitle": "hero_subtitle",
    "description": "hero_subtitle",
    "hero_image_url": "hero_image_url",
    "image_url": "hero_image_url",
    "image": "hero_image_url",
    "photo_url": "hero_image_url",
    "listing photo": "hero_image_url",
    "listing_photo": "hero_image_url",
    "img url": "hero_image_url",
    "img_url": "hero_image_url",
    "website_url": "website_url",
    "website": "website_url",
    "map_embed_url": "map_embed_url",
    "map_url": "map_embed_url",
    "google maps": "map_embed_url",
    "testimonial_quote": "testimonial_quote",
    "testimonial": "testimonial_quote",
    "testimonial_author": "testimonial_author",
    "author": "testimonial_author",
    "listing_plan": "listing_plan",
    "plan": "listing_plan",
    "timezone": "timezone",
    "upi_id": "upi_id",
    "upi": "upi_id",
    "slug": "slug",
}


def _norm_header(value: str) -> str:
    text = (value or "").strip().lower().replace("_", " ")
    return re.sub(r"\s+", " ", text)


# Normalized alias keys (underscores → spaces) for robust CSV header matching
_HEADER_LOOKUP = {_norm_header(k): v for k, v in HEADER_ALIASES.items()}


SAMPLE_ROWS = [
    {
        "name": "Sample Glow Salon",
        "industry_type": "salon",
        "public_phone": "9876543210",
        "public_email": "hello@sampleglow.example",
        "public_address": "Ring Road, Surat, Gujarat",
        "hero_title": "Hair & beauty near Ring Road",
        "hero_subtitle": "Walk-ins welcome. Bridal packages available.",
        "hero_image_url": "https://images.unsplash.com/photo-1560066984-138dadb4c035?w=640&q=45&fm=webp",
        "website_url": "https://example.com",
        "map_embed_url": "",
        "testimonial_quote": "Best salon in the area!",
        "testimonial_author": "Priya S.",
        "listing_plan": "free",
        "timezone": "Asia/Kolkata",
        "upi_id": "",
        "slug": "",
    },
    {
        "name": "Sample Care Dental",
        "industry_type": "dentist",
        "public_phone": "9123456780",
        "public_email": "care@sampledental.example",
        "public_address": "Adajan Gam, Surat",
        "hero_title": "Gentle dental care in Adajan",
        "hero_subtitle": "Cleaning, whitening, and family dentistry.",
        "hero_image_url": "https://images.unsplash.com/photo-1629909613654-28e377c37b09?w=640&q=45&fm=webp",
        "website_url": "",
        "map_embed_url": "",
        "testimonial_quote": "",
        "testimonial_author": "",
        "listing_plan": "pro",
        "timezone": "Asia/Kolkata",
        "upi_id": "",
        "slug": "",
    },
]

MAX_IMPORT_ROWS = 200
_URL_VALIDATOR = URLValidator()


@dataclass
class ImportRowResult:
    row_number: int
    name: str = ""
    ok: bool = False
    error: str = ""
    business_id: int | None = None
    slug: str = ""


@dataclass
class ImportResult:
    created: list[ImportRowResult] = field(default_factory=list)
    errors: list[ImportRowResult] = field(default_factory=list)
    skipped_empty: int = 0

    @property
    def created_count(self) -> int:
        return len(self.created)

    @property
    def error_count(self) -> int:
        return len(self.errors)


def map_headers(raw_headers: list[str]) -> dict[str, int]:
    """Map canonical column → index. Raises ValueError if name missing."""
    mapping: dict[str, int] = {}
    for idx, header in enumerate(raw_headers):
        key = _HEADER_LOOKUP.get(_norm_header(header))
        if key and key not in mapping:
            mapping[key] = idx
    if "name" not in mapping:
        raise ValueError(
            "Sheet must include a Name column (name / business name)."
        )
    return mapping


def _industry_lookup() -> dict[str, str]:
    """Map lowercase label or value → industry value."""
    lookup = {}
    for value, label in industry_choices_flat():
        lookup[value.lower()] = value
        lookup[label.lower()] = value
        lookup[label.lower().replace(" / ", "/")] = value
        lookup[label.lower().replace("/", " ")] = value
    return lookup


def _normalize_industry(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    lookup = _industry_lookup()
    hit = lookup.get(text.lower())
    if hit:
        return hit
    # fuzzy: strip punctuation
    compact = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    hit = lookup.get(compact)
    if hit:
        return hit
    valid = ", ".join(v for v, _ in industry_choices_flat()[:8]) + ", …"
    raise ValueError(f"Unknown category '{text}'. Use values like: {valid}")


def _normalize_plan(raw: str, *, allow_paid: bool) -> str:
    text = (raw or "").strip().lower() or "free"
    if text not in {"free", "pro", "premium"}:
        raise ValueError("listing_plan must be free, pro, or premium.")
    if not allow_paid and text != "free":
        return "free"
    return text


def _optional_url(raw: str, field_name: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    try:
        _URL_VALIDATOR(text)
    except ValidationError as exc:
        raise ValueError(f"Invalid {field_name}: {text}") from exc
    return text


def _cell(row: list, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def parse_row(row: list, header_map: dict[str, int], *, allow_paid_plans: bool) -> dict:
    def get(col: str) -> str:
        return _cell(row, header_map.get(col))

    name = get("name")
    if not name:
        return {}

    industry = _normalize_industry(get("industry_type") or "other")
    phone = get("public_phone")
    if phone:
        digits = "".join(ch for ch in phone if ch.isdigit())
        if len(digits) < 10:
            raise ValueError("Phone must have at least 10 digits.")

    data = {
        "name": name[:255],
        "industry_type": industry,
        "public_phone": phone[:20],
        "public_email": get("public_email")[:254],
        "public_address": get("public_address"),
        "hero_title": get("hero_title")[:255],
        "hero_subtitle": get("hero_subtitle"),
        "hero_image_url": _optional_url(get("hero_image_url"), "hero_image_url"),
        "website_url": _optional_url(get("website_url"), "website_url"),
        "map_embed_url": _optional_url(get("map_embed_url"), "map_embed_url"),
        "testimonial_quote": get("testimonial_quote"),
        "testimonial_author": get("testimonial_author")[:150],
        "listing_plan": _normalize_plan(get("listing_plan"), allow_paid=allow_paid_plans),
        "timezone": get("timezone") or "Asia/Kolkata",
        "upi_id": get("upi_id")[:50],
        "slug": get("slug")[:255],
    }
    email = data["public_email"]
    if email and "@" not in email:
        raise ValueError(f"Invalid email: {email}")
    return data


def rows_from_csv_text(text: str) -> tuple[list[str], list[list[str]]]:
    # Strip BOM
    if text.startswith("\ufeff"):
        text = text[1:]
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("File is empty.")
    return rows[0], rows[1:]


def rows_from_excel(file_obj: BinaryIO) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "Excel support requires openpyxl. Ask admin to install dependencies."
        ) from exc

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Excel sheet is empty.") from exc
    headers = ["" if c is None else str(c) for c in header]
    body = []
    for row in rows_iter:
        body.append(["" if c is None else str(c) for c in row])
    wb.close()
    return headers, body


def google_sheet_export_url(url: str) -> str:
    """Convert a Google Sheets share/edit URL to CSV export URL."""
    text = (url or "").strip()
    if not text:
        raise ValueError("Google Sheet URL is required.")
    if "export?format=csv" in text or "/export?" in text:
        return text

    parsed = urlparse(text)
    if "docs.google.com" not in parsed.netloc:
        raise ValueError("Paste a Google Sheets URL (docs.google.com/spreadsheets/…).")

    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", parsed.path)
    if not match:
        raise ValueError("Could not find spreadsheet ID in the URL.")
    sheet_id = match.group(1)

    gid = "0"
    qs = parse_qs(parsed.query)
    if "gid" in qs:
        gid = qs["gid"][0]
    elif parsed.fragment and "gid=" in parsed.fragment:
        frag_qs = parse_qs(parsed.fragment.replace("?", "&"))
        if "gid" in frag_qs:
            gid = frag_qs["gid"][0]

    return (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export"
        f"?format=csv&gid={gid}"
    )


def fetch_google_sheet_csv(url: str, *, timeout: int = 20) -> str:
    export_url = google_sheet_export_url(url)
    req = Request(
        export_url,
        headers={"User-Agent": "SuratBazarListingImport/1.0"},
    )
    try:
        with urlopen(req, timeout=timeout) as resp:
            content_type = (resp.headers.get("Content-Type") or "").lower()
            raw = resp.read()
    except HTTPError as exc:
        if exc.code in (401, 403):
            raise ValueError(
                "Sheet is not public. In Google Sheets: Share → Anyone with the link → Viewer."
            ) from exc
        raise ValueError(f"Could not download Google Sheet (HTTP {exc.code}).") from exc
    except URLError as exc:
        raise ValueError(f"Could not reach Google Sheets: {exc.reason}") from exc

    if "html" in content_type and b"<html" in raw[:200].lower():
        raise ValueError(
            "Google returned a login page. Make the sheet public "
            "(Anyone with the link can view), then try again."
        )
    return raw.decode("utf-8-sig", errors="replace")


def load_rows_from_upload(uploaded_file) -> tuple[list[str], list[list[str]]]:
    name = (getattr(uploaded_file, "name", "") or "").lower()
    raw = uploaded_file.read()
    if name.endswith((".xlsx", ".xlsm")):
        return rows_from_excel(io.BytesIO(raw))
    if name.endswith(".xls"):
        raise ValueError("Old .xls format not supported. Save as .xlsx or CSV.")
    # CSV / TSV / unknown → try decode
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    if name.endswith(".tsv") or "\t" in text.splitlines()[0]:
        reader = csv.reader(io.StringIO(text), delimiter="\t")
        rows = list(reader)
        if not rows:
            raise ValueError("File is empty.")
        return rows[0], rows[1:]
    return rows_from_csv_text(text)


def load_rows_from_source(*, uploaded_file=None, google_sheet_url: str = "") -> tuple[list[str], list[list[str]]]:
    if uploaded_file and getattr(uploaded_file, "size", None) != 0:
        # Empty file uploads sometimes still truthy; prefer URL if both empty
        name = getattr(uploaded_file, "name", "") or ""
        if name:
            return load_rows_from_upload(uploaded_file)
    if google_sheet_url.strip():
        text = fetch_google_sheet_csv(google_sheet_url)
        return rows_from_csv_text(text)
    raise ValueError("Upload a CSV/Excel file or paste a Google Sheet URL.")


def import_listing_rows(
    actor,
    headers: list[str],
    body_rows: list[list[str]],
    *,
    allow_paid_plans: bool = False,
) -> ImportResult:
    from catalog.services import ensure_default_hours, ensure_starter_categories
    from core.services import bust_directory_cache

    header_map = map_headers(headers)
    result = ImportResult()

    if len(body_rows) > MAX_IMPORT_ROWS:
        raise ValueError(f"Too many rows (max {MAX_IMPORT_ROWS}). Split the sheet.")

    created_any = False
    for offset, row in enumerate(body_rows, start=2):  # 1-indexed + header
        if not any(str(c or "").strip() for c in row):
            result.skipped_empty += 1
            continue
        try:
            data = parse_row(row, header_map, allow_paid_plans=allow_paid_plans)
            if not data:
                result.skipped_empty += 1
                continue
            business = _create_business_from_import(actor, data)
            ensure_starter_categories(business)
            ensure_default_hours(business)
            created_any = True
            result.created.append(
                ImportRowResult(
                    row_number=offset,
                    name=business.name,
                    ok=True,
                    business_id=business.pk,
                    slug=business.slug,
                )
            )
        except Exception as exc:  # noqa: BLE001 — collect per-row errors
            result.errors.append(
                ImportRowResult(
                    row_number=offset,
                    name=_cell(row, header_map.get("name")),
                    ok=False,
                    error=str(exc),
                )
            )

    if created_any:
        bust_directory_cache()
    return result


def _create_business_from_import(actor, data: dict) -> Business:
    slug = (data.get("slug") or "").strip()
    if slug and Business.objects.filter(slug=slug).exists():
        raise ValueError(f"Slug already exists: {slug}")

    setup_done = bool(
        data.get("public_phone")
        and data.get("public_address")
        and data.get("name")
    )

    business = Business(
        name=data["name"],
        industry_type=data["industry_type"],
        public_phone=data.get("public_phone") or "",
        public_email=data.get("public_email") or "",
        public_address=data.get("public_address") or "",
        hero_title=data.get("hero_title") or "",
        hero_subtitle=data.get("hero_subtitle") or "",
        hero_image_url=data.get("hero_image_url") or "",
        website_url=data.get("website_url") or "",
        map_embed_url=data.get("map_embed_url") or "",
        testimonial_quote=data.get("testimonial_quote") or "",
        testimonial_author=data.get("testimonial_author") or "",
        listing_plan=data.get("listing_plan") or "free",
        timezone=data.get("timezone") or "Asia/Kolkata",
        upi_id=data.get("upi_id") or "",
        created_by=actor if actor and actor.is_authenticated else None,
        profile_setup_completed=setup_done,
    )
    if slug:
        business.slug = slug
    business.save()

    # Partners who import own the listing
    if (
        actor
        and getattr(actor, "platform_role", None) == "marketing_partner"
        and actor.is_authenticated
    ):
        business.assigned_partners.add(actor)

    return business


def build_sample_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=IMPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in SAMPLE_ROWS:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def build_sample_xlsx() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"
    ws.append(IMPORT_COLUMNS)
    for row in SAMPLE_ROWS:
        ws.append([row.get(col, "") for col in IMPORT_COLUMNS])
    # Hint sheet
    tip = wb.create_sheet("Instructions")
    tip["A1"] = "How to import"
    tip["A2"] = "1. Fill rows on the Listings sheet (keep header row)."
    tip["A3"] = "2. industry_type: salon, dentist, optical, gym, restaurant, …"
    tip["A4"] = "3. listing_plan: free | pro | premium (partners import as free)."
    tip["A5"] = "4. hero_image_url: public HTTPS image URL (JPG/PNG/WebP)."
    tip["A6"] = "5. Google Sheets: File → Share → Anyone with link → Viewer, then paste URL."
    tip["A7"] = "6. Upload CSV/XLSX or paste Google Sheet URL in Partner / Admin import page."
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
